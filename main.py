import os

import requests
from telegram import Update
from telegram.ext import Updater, Filters, CommandHandler, MessageHandler, CallbackContext
from openpyxl import load_workbook
from datetime import datetime, timedelta, time
import pytz
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

# Your Telegram bot token from BotFather
TOKEN = os.getenv("ttoken")
sticker_id = "CAACAgUAAxkBAAETZyJiaC5D_9ue30Ae6sHQ1SogU5s7fgACGQEAAqP7yVR9wpCHJCufxiQE"
# Dictionary to store the day book data with datetime keys
date_party_dict = {}

# Get today's date in Indian Standard Time (IST)
indian_timezone = pytz.timezone('Asia/Kolkata')
today_date_ist = datetime.now(indian_timezone).date()


def start(update: Update, context: CallbackContext) -> None:
    update.message.reply_text('Hi! I am your DayBook Bot. Send me your DayBook.xlsx file.')


def handle_daybook(update: Update, context: CallbackContext) -> None:
    file_id = update.message.document.file_id
    file = context.bot.get_file(file_id)
    file.download('DayBook.xlsx')
    process_daybook(update, context)


def process_daybook(update: Update, context: CallbackContext) -> None:
    global date_party_dict

    # Load the Excel workbook
    excel_file_path = 'DayBook.xlsx'
    workbook = load_workbook(excel_file_path)

    # Select the active sheet or specify the sheet name
    sheet = workbook.active

    with open("bills.txt", 'r') as file:
        # Read all lines from the file
        lines = file.readlines()

        # Convert lines back to numbers
    read_numbers = []
    for line in lines:
        try:
            read_numbers.append(int(line.strip()))
        except ValueError:
            continue

    # Process the day book data starting from row 8
    for row in sheet.iter_rows(min_row=8, values_only=True):
        original_date = row[0]
        new_date = original_date + timedelta(days=15)

        new_date = new_date.date()
        if new_date <= today_date_ist:
            # Append party name to the list in the dictionary
            if int(row[3]) not in read_numbers and [
                row[0].strftime("%d-%b-%y"), row[1], row[3], row[4]] not in date_party_dict.get(new_date, []):
                date_party_dict.setdefault(new_date, []).append([row[0].strftime("%d-%b-%y"), row[1], row[3], row[4]])

    # Clean up old data
    for date, values in list(date_party_dict.items()):
        if date < (today_date_ist - timedelta(days=20)):
            del date_party_dict[date]
        new_sub_list = []
        for v in values:
            if int(v[2]) in read_numbers:
                continue
            new_sub_list.append(v)
        if new_sub_list:
            date_party_dict[date] = new_sub_list
        else:
            del date_party_dict[date]

    # Notify user about the update
    formatted_today_date = today_date_ist.strftime("%d-%b-%y")
    url = "https://api.telegram.org/bot{}/sendMessage?chat_id=781472777&text" \
          "={}".format(TOKEN, f"Data updated for {formatted_today_date}")
    requests.get(url)
    send_report(update, context)


def reply(update: Update, context: CallbackContext):
    user_input = update.message.text
    split_data = user_input.split()
    if split_data[0] == "Done":
        for i in split_data[1:]:
            try:
                bill_no = int(i)
                file_path = 'bills.txt'
                with open(file_path, 'r') as file:
                    # Read all lines from the file
                    lines = file.readlines()

                # Convert lines back to numbers
                read_numbers = [int(line.strip()) for line in lines]
                if bill_no in read_numbers:
                    update.message.reply_text(i + " is already updated")
                else:
                    with open(file_path, 'a') as file:
                        file.write(i + '\n')
                    update.message.reply_text(i + " is updated")

            except ValueError:
                update.message.reply_text(i + "is not a valid bill no")
        process_daybook(update, context)
    else:
        url = "https://api.telegram.org/bot{}/sendSticker?chat_id={}&sticker={}".format(
            TOKEN, update.message.chat.id, sticker_id)
        _ = requests.get(url)


def send_report(update: Update, context: CallbackContext):
    chat_id = 781472777
    formatted_today_date = today_date_ist.strftime("%d-%b-%y")
    if date_party_dict:
        report = ""
        for k, v in date_party_dict.items():
            report += k.strftime("%d-%b-%y ") + str(v)
            report += "\n\n"
        if context:
            context.bot.send_message(chat_id, report)
        else:
            url = "https://api.telegram.org/bot{}/sendMessage?chat_id=781472777&text" \
                  "={}".format(TOKEN, report)
            requests.get(url)
    else:
        if context:
            context.bot.send_message(chat_id, f"No data available for {formatted_today_date}.")
        else:
            url = "https://api.telegram.org/bot{}/sendMessage?chat_id=781472777&text" \
                  "={}".format(TOKEN, "No data")
            requests.get(url)


def unknown(update: Update, context: CallbackContext):
    update.message.reply_text(
        "Sorry '%s' is not a valid command" % update.message.text)


def main():
    updater = Updater(TOKEN)
    dp = updater.dispatcher

    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(MessageHandler(Filters.document, handle_daybook))
    dp.add_handler(CommandHandler("process", process_daybook))
    dp.add_handler(MessageHandler(Filters.text, reply))
    dp.add_handler(MessageHandler(
        Filters.command, unknown))
    updater.start_polling()
    scheduler = BackgroundScheduler(timezone=pytz.timezone('Asia/Kolkata'))
    scheduler.add_job(process_daybook, CronTrigger(hour=16, minute=31, second=0), id='daily_update', args=[None, None])
    scheduler.start()
    updater.idle()


if __name__ == '__main__':
    main()
