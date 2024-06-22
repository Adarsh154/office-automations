import os
from datetime import datetime, timedelta

import pytz
import requests
from openpyxl import load_workbook

# Your Telegram bot token from BotFather
TOKEN = "Token_here"


def process_daybook():
    date_party_dict = {}
    # Get today's date in Indian Standard Time (IST)
    indian_timezone = pytz.timezone('Asia/Kolkata')
    today_date_ist = datetime.now(indian_timezone).date()

    # Load the Excel workbook
    excel_file_path = 'DayBook.xlsx'
    workbook = load_workbook(excel_file_path)

    # Select the active sheet or specify the sheet name
    sheet = workbook.active
    # Process the day book data starting from row 8
    for row in sheet.iter_rows(min_row=8, values_only=True):
        original_date = row[0]
        new_date = original_date + timedelta(days=15)

        new_date = new_date.date()
        if new_date == today_date_ist:
            # Append party name to the list in the dictionary
            date_party_dict.setdefault(new_date, []).append(
                [row[0].strftime("%d-%b-%y"), row[1][:row[1].find("Rane")], row[3], row[4]])

    # Notify user about the update
    formatted_today_date = today_date_ist.strftime("%d-%b-%y")
    url = "https://api.telegram.org/bot{}/sendMessage?chat_id=781472777&text" \
          "={}".format(TOKEN, f"Data updated for {formatted_today_date}")
    requests.get(url)
    send_report(date_party_dict)


def send_report(date_party_dict):
    if date_party_dict:
        report = ""
        for k, v in date_party_dict.items():
            report += k.strftime("%d-%b-%y ") + str(v)
            report += "\n\n"
        report += "\n\n"
        url = "https://api.telegram.org/bot{}/sendMessage?chat_id=781472777&text" \
              "={}".format(TOKEN, report)
        requests.get(url)
    else:
        url = "https://api.telegram.org/bot{}/sendMessage?chat_id=<chatid>&text" \
              "={}".format(TOKEN, "No data")
        requests.get(url)


if __name__ == '__main__':
    process_daybook()
